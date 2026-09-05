from __future__ import annotations
import math
from pathlib import Path
import sys
import unittest
import zipfile
import openpyxl

ROOT=Path(__file__).resolve().parents[1]
sys.path.insert(0,str(ROOT/'scripts'))
from process_model import DATA, PRODUCTS, ELEMENTS, block_row, reference_result, samples, regulate, cleaning_events
from recalculate import Evaluator

PATH=ROOT/'deliverables/Стандартизированная_работа_И-ГЦ-02.xlsx'


class WorkbookTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.engine=Evaluator(PATH)
        cls.book=openpyxl.load_workbook(PATH)
        cls.cached=openpyxl.load_workbook(PATH,data_only=True)

    def evaluate(self,changes=None,check_errors=True):
        e=self.engine.calculate(changes)
        if check_errors:self.assertEqual(e.errors(),{},f'Formula errors for {changes}: {e.errors()}')
        return e

    def close(self,actual,expected):
        self.assertTrue(math.isclose(actual,expected,rel_tol=1e-9,abs_tol=1e-7),(actual,expected))

    def test_01_structure_and_protection(self):
        self.assertEqual(len(self.book.sheetnames),28)
        for name in ['Время такта','Лист наблюдений','Карта СР','Ручная работа','Периодическая работа',
                     'Производственная способность','ОТ ДО','ОТ ЦЕЛЬ','Диаграмма','Отчет об улучшении',
                     'Общий список улучшений','СОП','Критерии качества СР','Чек-лист СОП']:
            self.assertIn(name,self.book.sheetnames)
        with zipfile.ZipFile(PATH) as z:
            self.assertIsNone(z.testzip())
            self.assertFalse(any('externalLinks/' in p or 'vbaProject' in p for p in z.namelist()))
        # OOXML default is automatic when calcMode is absent (XlsxWriter).
        self.assertIn(self.book.calculation.calcMode,(None,'auto'))
        for sheet in self.book:
            self.assertTrue(sheet.protection.sheet,sheet.title)
            self.assertFalse(sheet.protection.password,sheet.title)
            for row in sheet:
                for c in row:
                    if c.data_type=='f':
                        self.assertTrue(c.protection.locked,(sheet.title,c.coordinate))
                        self.assertLess(len(c.value),8192)
                    elif c.fill.fgColor.rgb=='FFFFF2AC':
                        self.assertFalse(c.protection.locked,(sheet.title,c.coordinate))
                        self.assertIsNotNone(c.comment,(sheet.title,c.coordinate))
        self.assertEqual(self.cached['Начало']['C18'].value,0)

    def test_02_baseline_matches_independent_model(self):
        e=self.evaluate()
        for target in [False,True]:
            state='ЦЕЛЬ' if target else 'ДО';ref=reference_result(target)
            for cell,key in [('B30','cyclic'),('B32','periodic'),('B33','total'),('B34','mean_total'),
                             ('B35','available'),('B36','load'),('B38','active'),('B41','machine')]:
                self.close(e.get('Расчет '+state,cell),ref[key])
            for i in range(18):
                b=block_row(i);r=regulate(samples(i,target))
                for j in range(11):
                    self.close(e.get('Замеры '+state,f'R{b+5+j}'),r['adjustments'][j])
                    self.close(e.get('Замеры '+state,f'S{b+5+j}'),r['adjusted'][j])
                self.close(e.get('Замеры '+state,f'S{b+16}'),r['t_min'])
        self.close(e.get('Ввод','B30'),768.75)
        self.assertEqual(e.get('Контроль','C37'),0)
        self.assertEqual(e.get('Расчет ДО','B50'),'НЕ ВПИСЫВАЕТСЯ')
        self.assertEqual(e.get('Расчет ЦЕЛЬ','B50'),'ВПИСЫВАЕТСЯ')

    def test_03_change_shift_recalculates(self):
        e=self.evaluate({('Ввод','B17'):540})
        self.close(e.get('Ввод','B27'),28200)
        self.close(e.get('Ввод','B30'),28200/32)
        self.close(e.get('Диаграмма','H8'),28200/32)
        self.close(e.get('Расчет ДО','B36'),reference_result()['total']/28200)

    def test_04_change_volume_and_round_cleaning(self):
        q=[p['demo_batches']*p['max_kg'] for p in PRODUCTS];q[0]=6
        e=self.evaluate({('Ассортимент','F8'):6})
        self.assertEqual(e.get('Ввод','B28'),34)
        self.assertEqual(e.get('Периодическая работа','B20'),9)
        self.assertEqual(e.get('Периодическая работа','O13'),5)
        self.close(e.get('Расчет ДО','B33'),reference_result(quantities=q)['total'])
        self.close(e.get('Ввод','B30'),24600/34)

    def test_05_yield_and_mass_balance(self):
        e=self.evaluate({('Ассортимент','H8'):.8})
        self.close(e.get('Ассортимент','K8'),3.75)
        self.close(e.get('Ассортимент','L8'),3.75/.95)
        self.close(e.get('Ассортимент','O8'),3)
        self.close(e.get('Ассортимент','P8'),.75)
        y=[1]*18;y[0]=.8
        self.close(e.get('Расчет ДО','B33'),reference_result(yields_wash=y)['total'])

    def test_06_excludes_whole_cycle(self):
        e=self.evaluate({('Замеры ДО','E10'):0})
        reference=regulate(samples(0),[False]+[True]*9)
        self.assertEqual(e.get('Замеры ДО','T24'),9)
        self.assertEqual(e.get('Лист наблюдений','C8'),0)
        for j in range(11):
            self.close(e.get('Замеры ДО',f'S{12+j}'),reference['adjusted'][j])
        self.close(e.get('ОТ ДО','H21'),reference['t_min'])
        self.assertEqual(e.get('Контроль','C20'),'ТРЕБУЕТ ВНИМАНИЯ')

    def test_07_missing_element_excludes_whole_cycle(self):
        e=self.evaluate({('Замеры ДО','E12'):None})
        self.assertEqual(e.get('Замеры ДО','T24'),9)
        self.assertEqual(e.get('Лист наблюдений','C8'),0)
        self.close(e.get('ОТ ДО','H21'),regulate(samples(0),[False]+[True]*9)['t_min'])

    def test_08_select_product(self):
        e=self.evaluate({('Ввод','B12'):'Томат',('Ввод','B13'):'ЦЕЛЬ'})
        self.assertEqual(e.get('Ввод','B29'),4)
        self.close(e.get('СОП','F16'),180)
        self.assertIn('«СО»',e.get('СОП','C17'))
        self.close(e.get('Лист наблюдений','B20'),regulate(samples(3,True))['t_min'])
        self.close(e.get('ОТ ЦЕЛЬ','H21'),regulate(samples(3,True))['t_min'])

    def test_09_zero_demand(self):
        e=self.evaluate({('Ассортимент',f'F{r}'):0 for r in range(8,26)})
        self.assertEqual(e.get('Ввод','B28'),0)
        self.assertEqual(e.get('Ввод','B30'),'')
        self.assertEqual(e.get('Периодическая работа','B20'),0)
        self.assertEqual(e.get('Расчет ДО','B33'),0)
        self.assertEqual(e.get('Расчет ДО','B49'),'НЕТ РАСЧЁТА')

    def test_10_invalid_catalog_input(self):
        for change in [{('Ассортимент','H8'):0},{('Ассортимент','I8'):99},{('Ассортимент','F8'):-1}]:
            e=self.evaluate(change)
            self.assertNotEqual(e.get('Ассортимент','S8'),'ОК')
            self.assertEqual(e.get('Ассортимент','O8'),'')
            self.assertEqual(e.get('Ввод','B30'),'')
            self.assertGreater(e.get('Контроль','C37'),0)

    def test_11_invalid_fund_and_oee(self):
        e=self.evaluate({('Ввод','B17'):30})
        self.assertEqual(e.get('Ввод','B27'),'')
        self.assertEqual(e.get('Ввод','B30'),'')
        e=self.evaluate({('Показатели','B20'):0})
        self.assertEqual(e.get('Показатели','B37'),'')

    def test_12_machine_capacity_is_exact_not_rounded_up(self):
        for initial,closing in [(0,0),(0,1),(1,1),(3,0),(3,1)]:
            e=self.evaluate({('Ввод','B25'):initial,('Ввод','B26'):closing})
            available=e.get('Ввод','B27')
            for r in range(8,26):
                a=e.get('Производственная способность',f'F{r}')
                sanitary=e.get('Производственная способность',f'H{r}')
                capacity=e.get('Производственная способность',f'I{r}')
                def duration(n):return n*a+sum(cleaning_events(n,initial,bool(closing)))*sanitary
                self.assertLessEqual(duration(int(capacity)),available+1e-6)
                self.assertGreater(duration(int(capacity)+1),available-1e-6)

    def test_13_audit_does_not_claim_completion(self):
        e=self.evaluate()
        self.assertEqual(e.get('Чек-лист СОП','D31'),'')
        self.assertIn('НЕ ЗАВЕРШЕНА',e.get('Чек-лист СОП','B33'))
        self.assertEqual(e.get('Отчет об улучшении','E30'),'')
        self.assertEqual(e.get('Критерии качества СР','H16'),'Оценка не завершена')
        changes={('Чек-лист СОП',f'C{r}'):'Выполнено' for r in range(8,27)}
        e=self.evaluate(changes)
        self.assertEqual(e.get('Чек-лист СОП','D31'),19)
        self.assertIn('система работает',e.get('Чек-лист СОП','B33'))

    def test_14_fallback_regulation_no_variation_in_manual(self):
        changes={}
        s=samples(0)
        for i,el in enumerate(ELEMENTS):
            if el['kind']=='Ручная':
                s[i]=[s[i][0]]*10
                for j,c in enumerate('EFGHIJKLMN'):changes['Замеры ДО',f'{c}{12+i}']=s[i][j]
        e=self.evaluate(changes)
        reference=regulate(s)
        self.close(e.get('Замеры ДО','S23'),reference['t_min'])
        self.close(e.get('Замеры ДО','V24'),0)
        for i in range(11):self.close(e.get('Замеры ДО',f'R{12+i}'),reference['adjustments'][i])


    def test_15_no_complete_cycles_or_periodic_observation(self):
        e=self.evaluate({('Замеры ДО',f'{c}10'):0 for c in 'EFGHIJKLMN'})
        self.assertEqual(e.get('Расчет ДО','B33'),'')
        self.assertEqual(e.get('Общий список улучшений','F8'),'')
        e=self.evaluate({('Периодическая работа','E8'):None})
        self.assertEqual(e.get('Расчет ДО','B33'),'')
        self.assertEqual(e.get('Расчет ДО','M27'),'')
        self.assertGreater(e.get('Контроль','C37'),0)

    def test_16_sanitation_not_double_counted(self):
        changes={('Периодическая работа','E8'):90,('Периодическая работа','F8'):95,
                 ('Периодическая работа','G8'):92}
        e=self.evaluate(changes)
        self.assertEqual(e.get('Ввод','B27'),24600)
        for state in ['ДО','ЦЕЛЬ']:
            baseline=reference_result(state=='ЦЕЛЬ')
            self.close(e.get('Расчет '+state,'B32'),baseline['periodic']+8*30)
            self.close(e.get('Расчет '+state,'B33'),baseline['total']+8*30)
        self.assertEqual(e.get('Периодическая работа','J8'),90)

    def test_17_aggregate_resources_do_not_change_takt(self):
        e=self.evaluate({('Ввод','B22'):2,('Ввод','B23'):2})
        self.close(e.get('Ввод','B30'),768.75)
        ref=reference_result()
        self.close(e.get('Расчет ДО','B56'),ref['machine']/24600/2)
        self.close(e.get('Расчет ДО','B57'),ref['total']/24600/2)
        self.assertEqual(e.get('Контроль','C17'),'ТРЕБУЕТ ВНИМАНИЯ')

if __name__=='__main__':unittest.main(verbosity=2)
